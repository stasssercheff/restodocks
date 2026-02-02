#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Связывает цены из листа «Продукты_цены» со всеми карточками и листами.
Добавляет формулы пересчёта стоимости с учётом отхода и ужарки.

Использование: python ttk_link_prices.py /Users/masurfsker/Desktop/ТТК.xlsx
Результат сохраняется в тот же файл (создаётся копия _linked.xlsx).
"""
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Установи openpyxl: pip install openpyxl")
    sys.exit(1)

PRICE_SHEET = "Продукты_цены"
PRICE_NAME_COL = 2   # B - Наименование
PRICE_COST_COL = 3  # C - Стоимость за кг/шт
PRICE_RANGE = "Продукты_цены!$B:$C"


def set_pf_formulas(ws):
    """ПФ: колонка H (Цена за кг/шт) = VLOOKUP по Продукт (B)."""
    for r in range(3, ws.max_row + 1):
        product = ws.cell(r, 2).value
        if not product or not str(product).strip():
            continue
        ws.cell(r, 8).value = f"=IFNA(VLOOKUP(B{r},{PRICE_RANGE},2,FALSE),0)"
    print("  ПФ: цены привязаны к Продукты_цены")


def add_card_columns_and_formulas(ws, sheet_name):
    """Карточки: добавляем колонки 7–10 (Отход %, Ужарка %, Цена, Стоимость)."""
    headers = ("Отход %", "Ужарка %", "Цена", "Стоимость")
    seen_headers = set()
    for r in range(1, ws.max_row + 1):
        a_val = ws.cell(r, 1).value
        b_val = ws.cell(r, 2).value
        # Строка заголовков блока: A = №, B = Ингридиент
        if b_val == "Ингридиент" and (a_val == "№" or a_val == "#"):
            for c, h in enumerate(headers, start=7):
                ws.cell(r, c).value = h
            seen_headers.add(r)
        # Строка данных: A = число, B = название ингредиента
        elif isinstance(a_val, (int, float)) and a_val == int(a_val) and b_val and str(b_val).strip():
            # Отход % и Ужарка % — пусто или 0 (пользователь заполнит)
            if ws.cell(r, 7).value is None:
                ws.cell(r, 7).value = 0
            if ws.cell(r, 8).value is None:
                ws.cell(r, 8).value = 0
            # Цена из Продукты_цены (0 если продукт не найден)
            ws.cell(r, 9).value = f"=IFNA(VLOOKUP(B{r},{PRICE_RANGE},2,FALSE),0)"
            # Стоимость: (Шт/гр с учётом отхода и ужарки) * Цена / 1000
            ws.cell(r, 10).value = f"=D{r}/((1-G{r}/100)*(1-H{r}/100))*I{r}/1000"
    print(f"  {sheet_name}: добавлены колонки Отход %, Ужарка %, Цена, Стоимость и формулы")


def set_sendvichi_formulas(ws):
    """Сендвичи: колонка D (Цена) = VLOOKUP, E (стоимость) = C*D/1000 (граммы)."""
    for r in range(3, ws.max_row + 1):
        product = ws.cell(r, 2).value
        qty = ws.cell(r, 3).value
        if product is None or (qty is None and ws.cell(r, 4).value is None):
            continue
        if not str(product).strip():
            continue
        ws.cell(r, 4).value = f"=IFNA(VLOOKUP(B{r},{PRICE_RANGE},2,FALSE),0)"
        ws.cell(r, 5).value = f"=C{r}*D{r}/1000"
    print("  Сендвичи: цены и стоимость привязаны")


def set_novogodnee_formulas(ws):
    """Новогоднее меню: колонка D (Цена) = VLOOKUP, E уже =C*D/1000."""
    for r in range(3, ws.max_row + 1):
        product = ws.cell(r, 2).value
        if not product or not str(product).strip():
            continue
        ws.cell(r, 4).value = f"=IFNA(VLOOKUP(B{r},{PRICE_RANGE},2,FALSE),0)"
        if ws.cell(r, 5).value is None or not str(ws.cell(r, 5).value).startswith("="):
            ws.cell(r, 5).value = f"=C{r}*D{r}/1000"
    print("  Новогоднее меню 24-25: цены привязаны")


def main():
    path = Path("/Users/masurfsker/Desktop/ТТК.xlsx")
    if len(sys.argv) > 1:
        path = Path(sys.argv[1])
    if not path.exists():
        print(f"Файл не найден: {path}")
        sys.exit(1)

    out_path = path.parent / f"{path.stem}_linked.xlsx"
    wb = openpyxl.load_workbook(path, data_only=False)

    if PRICE_SHEET not in wb.sheetnames:
        print(f"Лист «{PRICE_SHEET}» не найден.")
        sys.exit(1)

    # ПФ
    if "ПФ" in wb.sheetnames:
        set_pf_formulas(wb["ПФ"])

    # Карточки Кухня и десерты
    if "Карточки Кухня" in wb.sheetnames:
        add_card_columns_and_formulas(wb["Карточки Кухня"], "Карточки Кухня")
    if "Карточки десерты" in wb.sheetnames:
        add_card_columns_and_formulas(wb["Карточки десерты"], "Карточки десерты")

    # Сендвичи, Новогоднее
    if "Сендвичи" in wb.sheetnames:
        set_sendvichi_formulas(wb["Сендвичи"])
    if "Новогоднее меню 24-25" in wb.sheetnames:
        set_novogodnee_formulas(wb["Новогоднее меню 24-25"])

    wb.save(out_path)
    print(f"\nГотово. Сохранено: {out_path}")


if __name__ == "__main__":
    main()
