#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Добавляет в лист «Продукты_цены» все ингредиенты из карточек и ПФ,
которых там ещё нет. Цены остаются пустыми — заполняешь один раз в Продукты_цены,
и они подтягиваются везде.

Использование: python ttk_add_missing_products.py /Users/masurfsker/Desktop/ТТК_linked.xlsx
Сохраняет в тот же каталог: имя_файла_full.xlsx
"""
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Установи openpyxl: pip install openpyxl")
    sys.exit(1)

PRICE_SHEET = "Продукты_цены"
SKIP_NAMES = {"продукт", "ингридиент", "итого", "количество", "наименование", "описание", "description"}


def collect_ingredients(wb):
    """Собирает все уникальные названия ингредиентов из карточек и ПФ."""
    ingredients = set()
    for sheet_name in ("ПФ", "Карточки Кухня", "Карточки десерты", "Сендвичи", "Новогоднее меню 24-25"):
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for r in range(2, min(ws.max_row + 1, 1500)):
            val = ws.cell(r, 2).value
            if not val or not str(val).strip():
                continue
            s = str(val).strip()
            if s.lower() in SKIP_NAMES:
                continue
            ingredients.add(s)
    return ingredients


def main():
    path = Path("/Users/masurfsker/Desktop/ТТК_linked.xlsx")
    if len(sys.argv) > 1:
        path = Path(sys.argv[1])
    if not path.exists():
        print(f"Файл не найден: {path}")
        sys.exit(1)

    out_path = path.parent / f"{path.stem}_full.xlsx"
    wb = openpyxl.load_workbook(path, data_only=False)

    if PRICE_SHEET not in wb.sheetnames:
        print(f"Лист «{PRICE_SHEET}» не найден.")
        sys.exit(1)

    ws = wb[PRICE_SHEET]
    existing = set()
    last_row = 2
    for r in range(3, ws.max_row + 1):
        name = ws.cell(r, 2).value
        if not name or not str(name).strip():
            break
        existing.add(str(name).strip())
        last_row = r

    ingredients = collect_ingredients(wb)
    missing = sorted(ingredients - existing)

    if not missing:
        print("Все ингредиенты уже есть в Продукты_цены. Ничего не добавлено.")
        wb.save(out_path)
        print(f"Файл сохранён: {out_path}")
        return

    next_row = last_row + 1
    next_num = next_row - 2  # № в колонке A
    for i, name in enumerate(missing):
        r = next_row + i
        ws.cell(r, 1).value = next_num + i
        ws.cell(r, 2).value = name
        ws.cell(r, 3).value = None  # Стоимость — заполнишь вручную
        ws.cell(r, 4).value = None  # Поставщик
        # Колонки 5–8 КБЖУ (если есть) — оставляем пусто, при следующем ttk_add_kbju подтянутся

    print(f"В «Продукты_цены» добавлено {len(missing)} позиций (без цены).")
    print("Заполни цены в листе «Продукты_цены» — они подтянутся во все карточки и ПФ.")
    wb.save(out_path)
    print(f"Сохранено: {out_path}")


if __name__ == "__main__":
    main()
