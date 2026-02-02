#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Добавляет КБЖУ (калории, белки, жиры, углеводы на 100 г) из JSON в таблицу ТТК:
— в лист «Продукты_цены» колонки Ккал, Белки, Жиры, Углеводы;
— в карточках — расчёт КБЖУ на ингредиент и итого на блюдо.

Использование: python ttk_add_kbju.py /Users/masurfsker/Desktop/ТТК_linked.xlsx
Или: python ttk_add_kbju.py /Users/masurfsker/Desktop/ТТК.xlsx
Сохраняет в тот же каталог: имя_файла_kbju.xlsx
"""
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Установи openpyxl: pip install openpyxl")
    sys.exit(1)

PRICE_SHEET = "Продукты_цены"
ROOT = Path(__file__).resolve().parent.parent
SCRIPT_DIR = Path(__file__).resolve().parent
EXTENDED_JSON = ROOT / "system_products_extended.json"
STARTER_JSON = ROOT / "Restodocks" / "starter_catalog.json"
OVERRIDE_JSON = SCRIPT_DIR / "ttk_kbju_override.json"


def load_kbju_dict():
    """Собирает словарь: нормализованное имя продукта -> {calories, protein, fat, carbs} (на 100 г)."""
    kbju = {}
    for path in (EXTENDED_JSON, STARTER_JSON):
        if not path.exists():
            continue
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        for item in data:
            names = []
            if item.get("name"):
                names.append(str(item["name"]).strip())
            nr = (item.get("names") or {}).get("ru")
            if nr and str(nr).strip() not in names:
                names.append(str(nr).strip())
            cal = item.get("calories") or item.get("kcal")
            if cal is None:
                continue
            row_val = {
                "calories": float(cal) if cal is not None else 0,
                "protein": float(item.get("protein", 0) or 0),
                "fat": float(item.get("fat", 0) or 0),
                "carbs": float(item.get("carbs", 0) or 0),
            }
            for name in names:
                if not name:
                    continue
                key = name.lower().strip()
                if key not in kbju:
                    kbju[key] = row_val
    return kbju


def normalize_name(s):
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def load_override_dict():
    """Ручной маппинг: название в ТТК -> КБЖУ (для продуктов, которых нет в JSON)."""
    if not OVERRIDE_JSON.exists():
        return {}
    with open(OVERRIDE_JSON, "r", encoding="utf-8") as f:
        data = json.load(f)
    return {normalize_name(k): v for k, v in data.items()}


def find_kbju(kbju_dict, product_name, override_dict=None):
    """Ищем КБЖУ: сначала ручной маппинг, потом JSON."""
    if not product_name or not str(product_name).strip():
        return None
    name = normalize_name(product_name)
    raw = str(product_name).strip()
    if override_dict and name in override_dict:
        return override_dict[name]
    if name in kbju_dict:
        return kbju_dict[name]
    without_pf = name.replace("пф ", "").strip()
    if without_pf in kbju_dict:
        return kbju_dict[without_pf]
    words = name.split()
    if words:
        first = words[0]
        if first in kbju_dict:
            return kbju_dict[first]
        # Префикс: «анчоус» → «анчоусы», «говядина» → «говядина вырезка»
        for key in kbju_dict:
            if key.startswith(first) or first in key:
                return kbju_dict[key]
        for key in kbju_dict:
            if key.startswith(name) or name.startswith(key):
                return kbju_dict[key]
    return None


def add_kbju_to_produkty_tseny(ws, kbju_dict, override_dict=None):
    """Продукты_цены: добавляем колонки 5–8 (Ккал, Белки, Жиры, Углеводы на 100 г)."""
    ws.cell(2, 5).value = "Ккал (100г)"
    ws.cell(2, 6).value = "Белки"
    ws.cell(2, 7).value = "Жиры"
    ws.cell(2, 8).value = "Углеводы"
    filled = 0
    for r in range(3, ws.max_row + 1):
        name = ws.cell(r, 2).value
        if not name or not str(name).strip():
            continue
        row_kbju = find_kbju(kbju_dict, name, override_dict)
        if row_kbju:
            ws.cell(r, 5).value = row_kbju["calories"]
            ws.cell(r, 6).value = row_kbju["protein"]
            ws.cell(r, 7).value = row_kbju["fat"]
            ws.cell(r, 8).value = row_kbju["carbs"]
            filled += 1
    print(f"  Продукты_цены: КБЖУ заполнены для {filled} продуктов (на 100 г)")


def add_kbju_to_cards(ws, sheet_name):
    """Карточки: колонки 11–14 — Ккал, Белки, Жиры, Углеводы на ингредиент; итого по блюду."""
    PRICE_RANGE = "Продукты_цены!$B:$H"  # B=имя, C=цена, D=Ккал, E=Белки, F=Жиры, G=Углеводы
    headers = ("Ккал", "Белки", "Жиры", "Углеводы")
    blocks = []  # (header_row, first_data_row, last_data_row)

    for r in range(1, ws.max_row + 1):
        a_val = ws.cell(r, 1).value
        b_val = ws.cell(r, 2).value
        if b_val == "Ингридиент" and (a_val == "№" or a_val == "#"):
            for c, h in enumerate(headers, start=11):
                ws.cell(r, c).value = h
            block_start = r + 1
            block_end = r + 1
            while block_end <= ws.max_row:
                aa = ws.cell(block_end, 1).value
                bb = ws.cell(block_end, 2).value
                if aa is None and bb is None:
                    break
                if bb == "Ингридиент" and (aa == "№" or aa == "#"):
                    break
                if isinstance(aa, (int, float)) and aa == int(aa) and bb and str(bb).strip():
                    block_end += 1
                else:
                    block_end += 1
            if block_end > block_start:
                blocks.append((r, block_start, block_end - 1))
        elif isinstance(a_val, (int, float)) and a_val == int(a_val) and b_val and str(b_val).strip():
            # КБЖУ на ингредиент: (Шт/гр/100)*значение на 100г; IFNA — 0 при отсутствии продукта
            for i, col in enumerate(range(11, 15)):
                ws.cell(r, col).value = f"=D{r}/100*IFNA(VLOOKUP(B{r},{PRICE_RANGE},{i+3},FALSE),0)"

    # Итого по блюду: вставляем строку после последней строки данных каждого блока
    offset = 0
    for header_row, start, end in blocks:
        insert_at = end + 1 + offset
        ws.insert_rows(insert_at)
        ws.cell(insert_at, 2).value = "Итого КБЖУ"
        for c, col in enumerate(range(11, 15)):
            ws.cell(insert_at, col).value = f"=SUM({openpyxl.utils.get_column_letter(col)}{start+offset}:{openpyxl.utils.get_column_letter(col)}{end+offset})"
        offset += 1

    print(f"  {sheet_name}: добавлены Ккал, Белки, Жиры, Углеводы на ингредиент и итого на блюдо")


def add_kbju_to_pf(ws):
    """ПФ: колонки 14–17 — Ккал, Белки, Жиры, Углеводы на строку (по Брутто)."""
    PRICE_RANGE = "Продукты_цены!$B:$H"
    ws.cell(2, 14).value = "Ккал"
    ws.cell(2, 15).value = "Белки"
    ws.cell(2, 16).value = "Жиры"
    ws.cell(2, 17).value = "Углеводы"
    for r in range(3, ws.max_row + 1):
        if not ws.cell(r, 2).value:
            continue
        for i, col in enumerate(range(14, 18)):
            ws.cell(r, col).value = f"=C{r}/100*IFNA(VLOOKUP(B{r},{PRICE_RANGE},{i+3},FALSE),0)"
    print("  ПФ: добавлены Ккал, Белки, Жиры, Углеводы на строку")


def main():
    path = Path("/Users/masurfsker/Desktop/ТТК_linked.xlsx")
    if len(sys.argv) > 1:
        path = Path(sys.argv[1])
    if not path.exists():
        print(f"Файл не найден: {path}")
        sys.exit(1)

    out_path = path.parent / f"{path.stem}_kbju.xlsx"
    kbju_dict = load_kbju_dict()
    override_dict = load_override_dict()
    print(f"Загружено КБЖУ: {len(kbju_dict)} из JSON, {len(override_dict)} из ручного маппинга")

    wb = openpyxl.load_workbook(path, data_only=False)

    if PRICE_SHEET not in wb.sheetnames:
        print(f"Лист «{PRICE_SHEET}» не найден.")
        sys.exit(1)

    add_kbju_to_produkty_tseny(wb[PRICE_SHEET], kbju_dict, override_dict)

    if "ПФ" in wb.sheetnames:
        add_kbju_to_pf(wb["ПФ"])

    for sheet_name in ("Карточки Кухня", "Карточки десерты"):
        if sheet_name in wb.sheetnames:
            add_kbju_to_cards(wb[sheet_name], sheet_name)

    wb.save(out_path)
    print(f"\nГотово. Сохранено: {out_path}")


if __name__ == "__main__":
    main()
